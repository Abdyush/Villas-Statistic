import os
import sys
import subprocess
import pickle
sys.path.append(os.getcwd())

import flet as ft
import openpyxl
from data.collection import butlers, perfect_new_month_statistic, fill_dict, clean_butlers
from data.data_output import new_fill_first_table, new_fill_second_table
from data.processing import new_prepare_dict, coef_new_count_totals
from src.copy_file import copy_file_to_folder
from src.background import prepare_lists_butlers
from src.resource import resource_path

 # Переменные для хранения путей к файлам
file_path_1 = ""
file_path_2 = ""
file_name_1 = ''
file_name_2 = ''
paths = {'one': '',
         'two': ''}

def main(page: ft.Page):
    page.theme_mode = ft.ThemeMode.DARK
    page.title = 'Villas Statistic'

#------------------------------------------------------ПЕРВАЯ СТРАНИЦА---------------------------------------------------------------------
    def first_page():
        page.clean()
            
        # Функция получения пути первого файла
        def return_first_file(e: ft.FilePickerResultEvent): 
            global file_path_1  # Используем global для изменения переменной в области видимости
            global file_name_1
            if e.files:  # Проверяем, что файл был выбран
                file_path_1 = e.files[0].path
                file_name_1 = e.files[0].name
                file_path_label_1.value = f"{file_name_1}"
                confirm_button_1.visible = True  # Показываем кнопку подтверждения
                confirm_button_1.disabled = False  
                cancel_button_1.visible = True  # Показываем кнопку отмены
                select_button_1.disabled = True  # Делаем кнопку выбора файла недоступной
            else:
                file_path_label_1.value = "Файл не выбран."
                confirm_button_1.visible = False  # Скрываем кнопку подтверждения
                cancel_button_1.visible = False  # Скрываем кнопку отмены
            file_path_label_1.update()
            filepicker_1.update()
            page.update()

        def return_second_file(e: ft.FilePickerResultEvent): 
            global file_path_2  # Используем global для изменения переменной в области видимости
            global file_name_2
            if e.files:  # Проверяем, что файл был выбран
                file_path_2 = e.files[0].path
                file_name_2 = e.files[0].name
                file_path_label_2.value = f"{file_name_2}"
                confirm_button_2.visible = True  # Показываем кнопку подтверждения
                confirm_button_2.disabled = False  
                cancel_button_2.visible = True  # Показываем кнопку отмены
                select_button_2.disabled = True
            else:
                file_path_label_2.value = "Файл не выбран."
                confirm_button_2.visible = False  # Скрываем кнопку подтверждения
                cancel_button_2.visible = False  # Скрываем кнопку отмены
            file_path_label_1.update()
            filepicker_2.update()
            page.update()
        
        
        def confirm_file_1(e):
            paths['one'] = file_path_1
            file_path_label_1.value = f"Подтвержден: {file_name_1}"
            confirm_button_1.disabled = True
            select_button_1.visible = False
            if paths['one'] != '' and paths['two'] != '':
                btn.disabled = False
            page.update()  
        
        def confirm_file_2(e):
            paths['two'] = file_path_2
            file_path_label_2.value = f"Подтвержден: {file_name_2}"
            confirm_button_2.disabled = True
            select_button_2.visible = False
            if paths['one'] != '' and paths['two'] != '':
                btn.disabled = False
            page.update()  

        def cancel_file_1(e):
            global file_path_1  # Объявляем переменную как глобальную
            global file_name_1
            nonlocal result
            file_path_1 = None  # Сбрасываем путь к файлу
            file_name_1 = ''
            paths['one'] = file_path_1
            file_path_label_1.value = ""  # Сбрасываем текст метки
            btn.visible = True
            confirm_button_1.visible = False  # Скрываем кнопку подтверждения
            cancel_button_1.visible = False  # Скрываем кнопку отмены
            select_button_1.visible = True
            select_button_1.disabled = False
            open_file_btn.visible = False
            result.value = ''
            if paths['one'] == None or paths['two'] == None:
                btn.disabled = True
            page.update()

        def cancel_file_2(e):
            global file_path_2  # Объявляем переменную как глобальную
            global file_name_2
            nonlocal result
            file_path_2 = None  # Сбрасываем путь к файлу
            file_name_2 = ''
            paths['two'] = file_path_2
            file_path_label_2.value = ""  # Сбрасываем текст метки
            btn.visible = True
            confirm_button_2.visible = False  # Скрываем кнопку подтверждения
            cancel_button_2.visible = False  # Скрываем кнопку отмены
            select_button_2.visible = True
            select_button_2.disabled = False
            open_file_btn.visible = False
            result.value = ''
            if paths['one'] == None or paths['two'] == None:
                btn.disabled = True
            page.update()
    
        def main_work(e, last_year, this_year):
            nonlocal result
            btn.disabled = True
            target_folder = 'distribution'
            
            try:
                # Путь к файлу эксель с распределением вилл 
                FILE_PATH = copy_file_to_folder(last_year, target_folder)   

                # Путь к файлу эксель с распределением вилл 
                FILE_24_PATH = copy_file_to_folder(this_year, target_folder)

                # Открываем файл
                file = openpyxl.open(FILE_PATH)
                file2 = openpyxl.open(FILE_24_PATH)

                loading1.visible = False
                loading2.visible = True
                page.update()

                # определяем список со всеми доступными страницами файла
                sheets = file.worksheets
                sheets2 = file2.worksheets

                list_sheet = [sheets, sheets2]

                other_comments = []

                # Добавляем батлеров из файла в словарь
                fill_dict(butlers)

                #Перебирая листы файла за каждый месяц, добавляем каждый лист в функцию получения статистики, таким образом, получаем статистику за год
                present_butlers, all_butlers = prepare_lists_butlers()
                for sheets in list_sheet:
                    for sheet in sheets:
                        perfect_new_month_statistic(sheets=sheets, sheet=sheet, sheet_index=sheets.index(sheet), 
                                                    sheets_index=list_sheet.index(sheets), 
                                                    present_butlers=present_butlers, all_butlers=all_butlers)

                new_prepare_dict(butlers)

                # Закрываем файлы
                file.close()  
                file2.close()                

                # Удаляем файлы
                os.remove(FILE_PATH)
                os.remove(FILE_24_PATH)

                generaly_dictionary = coef_new_count_totals(butlers)

                # open a pickle file
                filename = resource_path('butlers/selected_butlers.pk')

                with open(filename, 'wb') as fi:
                    # dump your data into the file
                    pickle.dump([], fi)

                # Путь к файлу со статистикой
                #STATISTIC_FILE_PATH = '/home/user/Рабочий стол/StudyProject/villas/Statistic/Butlers_statistic.xlsx'
                #STATISTIC_FILE_PATH = os.path.realpath('Statistic/Butlers_statistic.xlsx')
                STATISTIC_FILE_PATH = resource_path('Statistic/Butlers_statistic.xlsx')

                # Открываем файл cо статистикой
                file_stat = openpyxl.load_workbook(STATISTIC_FILE_PATH)

                # Удаляем все листы из файла
                all_sheet = file_stat.sheetnames
                for sh in all_sheet:
                    file_stat.remove(file_stat[sh])

                # Создаем листы "Все виллы", "Статистика"
                sheet = file_stat.create_sheet("Статистика")
                sheet1 = file_stat.create_sheet("Все виллы")

                # Заполняем вторую таблицу статистикой
                new_fill_second_table(sheet, generaly_dictionary)

                # Заполняем первую таблицу всеми данными по виллам
                new_fill_first_table(sheet1, dict(sorted(butlers.items())))

                # Выводим в терминал "Готово!", сохраняем и закрываем файл
                print('Готово!')
                clean_butlers(butlers)
                file_stat.save(STATISTIC_FILE_PATH)
                file_stat.close()
                cancel_button_1.disabled = False
                cancel_button_2.disabled = False
                btn.visible = False
                open_file_btn.visible = True
                open_file_btn.disabled = False

            except Exception as ex:
                # Отображаем сообщение об ошибке
                btn.visible = False
                open_file_btn.visible = False
                open_file_btn.update()
                result.color = "red"
                result.value = f"Ошибка: {str(ex)}"
                cancel_button_1.disabled = False
                cancel_button_2.disabled = False
                page.update()

            # Обновляем экран
            page.update()

        def open_stat_file(e):
            # Путь к файлу Excel
            #STATISTIC_FILE_PATH = os.path.realpath('Statistic/Butlers_statistic.xlsx')
            STATISTIC_FILE_PATH = resource_path('Statistic/Butlers_statistic.xlsx')

            # Открываем файл с помощью Excel (Windows)
            os.startfile(STATISTIC_FILE_PATH)

            # Или с помощью subprocess (для Linux)
            #subprocess.Popen(['xdg-open', STATISTIC_FILE_PATH])
            first_page()  


        def start_programm(e):
            loading1.visible = True
            cancel_button_1.disabled = True
            cancel_button_2.disabled = True
            btn.visible = False
            #loading.visible = True
            page.update()
            pb.visible = True
            pb.update()
            
            file_1 = paths['one']
            file_2 = paths['two']
            main_work(e, file_1, file_2)
            #pb.visible = False
            
            pb.visible = False
            loading2.visible = False
            pb.update()
            page.update()
            


        #---------------------------------------------------------Кнопки------------------------------------------------------------------
        # Кнопка перехода к настройкам
        settings_btn = ft.ElevatedButton("Настройки", icon=ft.Icons.SETTINGS, width=200, height=45, on_click=switch_to_second_page,
                                        style=ft.ButtonStyle(text_style=ft.TextStyle(size=17)))

        # Текст с результатом выполнения программы
        result = ft.Text(value="", theme_style=ft.TextThemeStyle.TITLE_MEDIUM, color="green")

        # Текстовые элементы для отображения статуса выбранного файла
        file_path_label_1 = ft.Text(value=file_name_1, expand=1, size=17)
        file_path_label_2 = ft.Text(value=file_name_2, expand=1, size=17)

        # Создаем filpickers - элементы управления для выбора файла
        filepicker_1 = ft.FilePicker(on_result=return_first_file)
        filepicker_2 = ft.FilePicker(on_result=return_second_file)

        page.overlay.append(filepicker_1)
        page.overlay.append(filepicker_2)
        
        # Кнопки для выбора файла
        select_button_1 = ft.ElevatedButton(text="Occupancy за прошлый год", icon=ft.Icons.UPLOAD, width=300, height=45, 
                                            on_click=lambda _: filepicker_1.pick_files(allowed_extensions=["xls", "xlsx", "xlsm", "xltx", "xltm"]),
                                            style=ft.ButtonStyle(text_style=ft.TextStyle(size=17)))
        select_button_2 = ft.ElevatedButton(text="Occupancy за текущий год", icon=ft.Icons.UPLOAD, width=300, height=45, 
                                            on_click=lambda _: filepicker_2.pick_files(allowed_extensions=["xls", "xlsx", "xlsm", "xltx", "xltm"]),
                                            style=ft.ButtonStyle(text_style=ft.TextStyle(size=17)))

        # Кнопки подтверждения
        confirm_button_1 = ft.IconButton(icon=ft.Icons.CHECK, width=70, height=45, on_click=confirm_file_1, visible=False)
        confirm_button_2 = ft.IconButton(icon=ft.Icons.CHECK, width=70, height=45, on_click=confirm_file_2, visible=False)

        # Кнопки отмены файла
        cancel_button_1 = ft.IconButton(icon=ft.Icons.CANCEL, width=100, height=45, on_click=cancel_file_1, visible=False)
        cancel_button_2 = ft.IconButton(icon=ft.Icons.CANCEL, width=100, height=45, on_click=cancel_file_2, visible=False)

        # Кнопка "Запуск"
        btn = ft.ElevatedButton(text='Запуск', icon=ft.Icons.DOUBLE_ARROW_OUTLINED, width=400, height=55, on_click=start_programm, 
                                style=ft.ButtonStyle(text_style=ft.TextStyle(size=17)), disabled=True)
        
        # Кнопка открытия файла со статистикой
        open_file_btn = ft.FilledButton(text='Открыть статистику', icon=ft.Icons.FILE_OPEN_ROUNDED, width=400, height=55, 
                                        on_click=open_stat_file, style=ft.ButtonStyle(text_style=ft.TextStyle(size=17)), visible=False)
        

        pb = ft.ProgressBar(width=400, visible=False)
        loading1 = ft.Text(value="Читаю файлы...", theme_style=ft.TextThemeStyle.TITLE_MEDIUM, visible=False)
        loading2 = ft.Text(value="Подвожу итоги...", theme_style=ft.TextThemeStyle.TITLE_MEDIUM, visible=False)


        # Проверяем были ли выбраны пути файлов, если да, то показываем кнопки подтверждения и отмены
        if file_path_label_1.value != "":
            confirm_button_1.visible = True  
            confirm_button_1.disabled = False  
            cancel_button_1.visible = True  
            select_button_1.disabled = True

        if file_path_label_2.value != "":
            confirm_button_2.visible = True  
            confirm_button_2.disabled = False  
            cancel_button_2.visible = True  
            select_button_2.disabled = True
            

        #------------------------------------------------Ряды и контейнеры-----------------------------------------------------------------
        # Первый ряд содержит кнопку перехода к настройкам
        first_row = ft.Row(controls=[ft.Column(controls=[ft.Container(height=10),
                                                         ft.Row(controls=[settings_btn, ft.Container(width=10)])])], 
                                                         alignment=ft.MainAxisAlignment.END)
        
        # Второй ряд содержит колонку, внутри которой первым рядом идет кнопка запуска, вторым - текст с результатом и кнопка открытия файла
        second_row = ft.Row(controls=[ft.Column(controls=[btn,
                                                          loading1,
                                                          loading2,
                                                          pb,
                                                          ft.Row(controls=[result, open_file_btn])],
                                                          alignment=ft.MainAxisAlignment.END)],
                                                height=320,
                                                alignment=ft.MainAxisAlignment.CENTER,
                                                )
        
        # Третий и четвертый ряд содержат кнопки: выбор файла, путь файла, отмена, подтвердить
        third_row = ft.Row(controls=[ft.Container(width=10), 
                                     select_button_1, file_path_label_1, cancel_button_1, confirm_button_1,
                                     ft.Container(width=10)])
        fourth_row = ft.Row(controls=[ft.Container(width=10), 
                                      select_button_2, file_path_label_2, cancel_button_2, confirm_button_2,
                                      ft.Container(width=10)])

        page.add(first_row, second_row, third_row, fourth_row)

        page.update()


#------------------------------------------------------ВТОРАЯ СТРАНИЦА---------------------------------------------------------------------
    def second_page():
        
        page.clean()  # Очищаем страницу
        page.scroll = ft.ScrollMode.ALWAYS
        
            # open a pickle file
        #filename1 = 'all_butlers.pk'
        #filename1 = os.path.realpath('butlers/all_butlers.pk')
        #filename2 = 'selected_butlers.pk'
        #filename2 = os.path.realpath('butlers/selected_butlers.pk')

        filename1 = resource_path('butlers/all_butlers.pk')
        filename2 = resource_path('butlers/selected_butlers.pk')


        list_textfields = []
        list_del_icons = []
        list_checkboxes = []

        # load your data back to memory when you need it
        with open(filename1, 'rb') as fi:
            all_butlers = pickle.load(fi)

        first_shift = all_butlers[0]
        second_shift = all_butlers[1]


        # Создаем контейнеры для сотрудников
        first_shift_list = ft.Column()
        second_shift_list = ft.Column()

        def populate_employee_lists(first_shift, second_shift):
            # Заполняем колонки начальными данными
            first_shift_list.controls.clear()
            second_shift_list.controls.clear()
            
            for name in first_shift:
                first_shift_list.controls.append(create_employee_row(name, first_shift_list))
            
            for name in second_shift:
                second_shift_list.controls.append(create_employee_row(name, second_shift_list))

            page.update()


        def add_employee(shift_column):
            # Добавляем нового сотрудника в выбранную смену
            employee_name = employee_input.value.strip()
            if employee_name:
                shift_column.controls.append(create_employee_row(employee_name, shift_column)) 
                employee_input.value = ""
                page.update()

        def create_employee_row(name, shift_column):
            # Создаем строку с именем сотрудника и чекбоксом
            checkbox = ft.Checkbox(label="", value=False, visible=False)
            text_field = ft.TextField(value=name, width=200, on_change=lambda e: update_employee_name(e, row), disabled=True, 
                                      color = ft.Colors.with_opacity(1, ft.Colors.PRIMARY), text_align=ft.TextAlign.CENTER, 
                                      text_style=ft.TextStyle(size=17))
            del_icon = ft.IconButton(ft.Icons.DIRECTIONS_RUN_OUTLINED, on_click=lambda e: delete_employee(row, shift_column), visible=False)
            row = ft.Row(controls=[text_field, checkbox, del_icon], alignment=ft.MainAxisAlignment.SPACE_BETWEEN)
            list_textfields.append(text_field)
            list_del_icons.append(del_icon)
            list_checkboxes.append(checkbox)

            return row
            
        def enable_text_fields():
            for text_field in list_textfields:
                text_field.disabled = False  # Устанавливаем свойство disabled в False
            for del_icon in list_del_icons:
                del_icon.visible = True
            employee_input.visible = True
            add_first_shift_button.visible = True
            add_second_shift_button.visible = True
            confirm_button.visible = True
            enable_button.visible = False
            select_individual_btn.visible = False
            cancel_button.visible = True

            page.update()


        def show_checkboxes():
            for checkbox in list_checkboxes:
                checkbox.visible = True
                
            confirm_selection_button.visible = True
            select_individual_btn.visible = False
            cancel_selection_btn.visible = True
            enable_button.visible = False

            page.update()
            
        def delete_employee(row, shift_column):
            # Удаляем сотрудника из списка
            shift_column.controls.remove(row)
            page.update()

        def update_employee_name(e, row):
            # Обновляем имя сотрудника в строке
            new_name = e.control.value.strip()
            if new_name:
                row.controls[0].value = new_name
                page.update()

        def cancel_selection():
            second_page()

            page.update()

        def confirm_changes(e):
            # Подтверждаем изменения и возвращаем обновленные списки
            updated_first_shift = [row.controls[0].value for row in first_shift_list.controls]
            updated_first_shift = sorted(updated_first_shift)
            updated_second_shift = [row.controls[0].value for row in second_shift_list.controls]
            updated_second_shift = sorted(updated_second_shift)
            all_but = [updated_first_shift, updated_second_shift]
            with open(filename1, 'wb') as fi:
            # dump your data into the file
                pickle.dump(all_but, fi)
            print("Обновленный список первой смены:", updated_first_shift)
            print("Обновленный список второй смены:", updated_second_shift)

            for text_field in list_textfields:
                text_field.disabled = True  # Устанавливаем свойство disabled в False
            for del_icon in list_del_icons:
                del_icon.visible = False
            employee_input.visible = False
            add_first_shift_button.visible = False
            add_second_shift_button.visible = False
            confirm_button.visible = False
            enable_button.visible = True
            select_individual_btn.visible = True
            cancel_button.visible = False

            page.update()

        def confirm_selection(e):
            nonlocal list_textfields 
            # Подтверждаем выбор сотрудников с галочками
            selected_first_shift = []
            selected_second_shift = []
            
            for row in first_shift_list.controls:
                checkbox = row.controls[1]  # Получаем чекбокс из строки
                if checkbox.value:
                    selected_first_shift.append(row.controls[0].value)  # Добавляем имя в список если чекбокс отмечен
            
            for row in second_shift_list.controls:
                checkbox = row.controls[1]  # Получаем чекбокс из строки
                if checkbox.value:
                    selected_second_shift.append(row.controls[0].value)  # Добавляем имя в список если чекбокс отмечен
            sel_but = [selected_first_shift, selected_second_shift]
            with open(filename2, 'wb') as fi:
            # dump your data into the file
                pickle.dump(sel_but, fi)
        
            list_all_selected = selected_first_shift + selected_second_shift

            confirm_selection_button.visible = False
            cancel_selection_btn.visible = False
            select_individual_btn.visible = True
            enable_button.visible = True
            for checkbox in list_checkboxes:
                checkbox.visible = False
            individual_stat.value = f'Программа посчитает статистику для выделенных батлеров'
            for butler in list_all_selected:
                for field in list_textfields:
                    if butler == field.value:
                        field.bgcolor=ft.Colors.LIGHT_BLUE_500
            enable_button.visible = False
            select_individual_btn.visible = False

            page.update()

        


        #---------------------------------------------------------Кнопки----------------------------------------------------------------------
        # Кнопка для возвращения к подсчету статистики
        count_stat_btn = ft.ElevatedButton("Подсчет статистики", icon=ft.Icons.FORMAT_LIST_NUMBERED, width=250, height=45, on_click=switch_to_first_page,
                                           style=ft.ButtonStyle(text_style=ft.TextStyle(size=17)))

        # Поле для ввода фамилии сотрудника
        employee_input = ft.TextField(label="Добавьте нового сотрудника", width=300, visible=False)
        
        # Кнопки добавления для каждой смены
        add_first_shift_button = ft.IconButton(ft.Icons.ADD, on_click=lambda e: add_employee(first_shift_list), visible=False)
        add_second_shift_button = ft.IconButton(ft.Icons.ADD, on_click=lambda e: add_employee(second_shift_list), visible=False)

        # Кнопка для включения всех TextField
        enable_button = ft.ElevatedButton(text="Внести изменения", width=310, height=50, icon=ft.Icons.EDIT_OUTLINED,
            on_click=lambda e: enable_text_fields(), style=ft.ButtonStyle(
            text_style=ft.TextStyle(size=17)) )
        
        # Кнопка подтверждения изменений
        confirm_button = ft.Button(text="Подтвердить изменения", width=250, height=45, icon=ft.Icons.CHECK, 
                                   on_click=confirm_changes, visible=False)
        # Кнопка отмены изменений
        cancel_button = ft.Button(text="Отменить изменения", width=250, height=45, icon=ft.Icons.CANCEL, 
                                  on_click=lambda e: cancel_selection(), visible=False)
        
        # Кнопка для выбора отдельных батлеров
        select_individual_btn = ft.Button(text='Выбрать отдельных батлеров', width=310, height=50, icon=ft.Icons.CHECK_BOX_OUTLINED, 
                                          on_click=lambda e: show_checkboxes(), style=ft.ButtonStyle(
                                          text_style=ft.TextStyle(size=17)))
        # Кнопка подтверждения выбора
        confirm_selection_button = ft.Button(text="Выбрать данных батлеров", width=250, height=45, icon=ft.Icons.CHECK,  
                                            on_click=confirm_selection,
                                            visible=False)
        # Кнопка отмены выбора отдельных батлеров 
        cancel_selection_btn = ft.Button(text="Отменить выбор", width=250, height=45, icon=ft.Icons.CANCEL, 
                                            on_click=lambda e:cancel_selection(),
                                            visible=False)
        # Текст с выбранными батлерами
        individual_stat = ft.Text("", size=18)


        # Заполняем начальные данные
        populate_employee_lists(first_shift, second_shift)


        #-------------------------------------------------------Ряды и контейнеры-------------------------------------------------------------
        # Первый ряд содержит кнопку возвращения на первую страницу в правом углу
        first_row = ft.Row(controls=[ft.Column(controls=[ft.Container(height=10),
                                                         ft.Row(controls=[count_stat_btn, ft.Container(width=10)])])], 
                                                         alignment=ft.MainAxisAlignment.END)
        
        # Второй ряд содержит текст "Список сотрудников по сменам"
        second_row = ft.Row(controls=[ft.Text("СПИСКИ БАТЛЕРОВ ПО СМЕНАМ", theme_style=ft.TextThemeStyle.TITLE_MEDIUM)], 
                            alignment=ft.MainAxisAlignment.CENTER)
        
        # Третий ряд содержит три колонки, по бокам - списки сотрудников, по середине - кнопки
        buttons_column = ft.Column(controls=[ft.Row(controls=[enable_button, cancel_button, confirm_button]),
                                             ft.Row(controls=[add_first_shift_button, employee_input, add_second_shift_button]),
                                             ft.Row(controls=[select_individual_btn, cancel_selection_btn, confirm_selection_button]),
                                             individual_stat],
                                             height=300,
                                             alignment=ft.MainAxisAlignment.END,
                                             horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                                             )
        third_row = ft.Row(controls=[ft.Column(controls=[ft.Text("Первая смена", theme_style=ft.TextThemeStyle.TITLE_MEDIUM), first_shift_list], 
                                               horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                                     buttons_column,
                                     ft.Column(controls=[ft.Text("Вторая смена", theme_style=ft.TextThemeStyle.TITLE_MEDIUM), second_shift_list], 
                                               horizontal_alignment=ft.CrossAxisAlignment.CENTER)],
                           alignment=ft.MainAxisAlignment.SPACE_EVENLY,
                           vertical_alignment=ft.CrossAxisAlignment.START)

        page.add(first_row, second_row, third_row)

     
    # Функция для переключения на первую страницу
    def switch_to_first_page(e):
        first_page()

    # Функция для переключения на вторую страницу
    def switch_to_second_page(e):
        second_page()

    # Запускаем приложение с первой страницы
    first_page()
    
if __name__ == "__main__":
    ft.app(main)