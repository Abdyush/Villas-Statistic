import os
import sys
import pickle
from openpyxl.cell import MergedCell
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from data.processing import detrmine_season
from src.resource import resource_path


# ------------------------ ПОДГОТОВИТЕЛЬНЫЕ ФУНКЦИИ -----------------------------------------

# Функция определяющая текущих батлеров, и формирующая список всех когда либо работающих
def prepare_lists_butlers():
    # open a pickle file
    #filename = 'all_time_butlers.pk'
    #filename1 = 'all_butlers.pk'
    #filename2 = 'selected_butlers.pk'

    #filename = os.path.realpath('butlers/all_time_butlers.pk')
    #filename1 = os.path.realpath('butlers/all_butlers.pk')
    #filename2 = os.path.realpath('butlers/selected_butlers.pk')

    filename = resource_path('butlers/all_time_butlers.pk')
    filename1 = resource_path('butlers/all_butlers.pk')
    filename2 = resource_path('butlers/selected_butlers.pk')

    # load your data back to memory when you need it
    with open(filename2, 'rb') as fi:
        sel_but = pickle.load(fi)

    if sel_but != []:
        present_butlers = sel_but[0]
        present_butlers.extend(sel_but[1])
    else:
        with open(filename1, 'rb') as fi:
            all_but = pickle.load(fi)
        present_butlers = all_but[0]
        present_butlers.extend(all_but[1])
        with open(filename, 'rb') as file:
            all_butlers = pickle.load(file)
            all_butlers = sorted(list(set(present_butlers) | set(all_butlers)))
        with open(filename, 'wb') as file:
            pickle.dump(all_butlers, file)
            print(all_butlers)

    with open(filename, 'rb') as fi:
        all_butlers = pickle.load(fi)

    return present_butlers, all_butlers

# Функция, определяющая диапазоны категорий вилл для отдельной страницы
def define_range(sheet_):
   for column in range(1, sheet_.max_column):
      if sheet_[1][column].fill.start_color.index == '00000000' and type(sheet_[1][column]) != MergedCell:
         column_letters = str(sheet_[1][column - 1].coordinate)[:-1]
         veg = sheet_['C5': column_letters + '16']
         vps = sheet_['C18': column_letters + '19']
         vig = sheet_['C21': column_letters + '23']
         fwv = sheet_['C33': column_letters + '48']
         pwv = sheet_['C50': column_letters + '53']
         categories = [(veg, 'VEG'), (vps, 'VPS'), (vig, 'VIG'), (fwv, 'FWV'), (pwv, 'PWV')]
         return categories


# Совершенная функция, определяющая диапазоны категорий вилл для отдельной страницы подходящая для обоих файлов
def perfect_define_range(sheet, index_sheet, index_sheets):
   # Определяем ряды начала и окончания диапазонов
   ranges = {'veg': ['4001', '4012'],
             'vps': ['5001', '5002'],
             'vig': ['5003', '5005'],
             'fwv': ['4013', '4028'],
             'pwv': ['5006', '5009']}
   if index_sheet == 11 and index_sheets == 0:
       col = 0
   else:
       col = 1
   for row in range(1, 70):
       for key, value in ranges.items():
           if str(sheet[row][col].value) in value:
               i = value.index(str(sheet[row][col].value))
               value[i] = row

   # Определяем колонки окончания диапазонов
   if index_sheet == 0 and index_sheets == 1:
      end_column_letter = 'BL'
   else:
      for column in range(1, sheet.max_column):
         if sheet[1][column].fill.start_color.index == '00000000' and type(sheet[1][column]) != MergedCell:
            end_column_letter = str(sheet[1][column - 1].coordinate)[:-1]
            break
         
   if index_sheet == 11 and index_sheets == 0:
      start_column_letter = 'B'
   else:
      start_column_letter = 'C'
         
   veg = sheet[start_column_letter + str(ranges['veg'][0]): end_column_letter + str(ranges['veg'][1])]
   vps = sheet[start_column_letter + str(ranges['vps'][0]): end_column_letter + str(ranges['vps'][1])]
   vig = sheet[start_column_letter + str(ranges['vig'][0]): end_column_letter + str(ranges['vig'][1])]
   fwv = sheet[start_column_letter + str(ranges['fwv'][0]): end_column_letter + str(ranges['fwv'][1])]
   pwv = sheet[start_column_letter + str(ranges['pwv'][0]): end_column_letter + str(ranges['pwv'][1])]

   categories = [(veg, 'VEG'), (vps, 'VPS'), (vig, 'VIG'), (fwv, 'FWV'), (pwv, 'PWV')]
   return categories

# Совершенная функция, определяющая диапазоны категорий вилл для отдельной страницы подходящая для обоих файлов
def perfect_perfect_define_range(sheet, index_sheet, index_sheets):
   diap = sheet.calculate_dimension()
   print(diap)
   print(type(diap))

   # Определяем ряды начала и окончания диапазонов
   ranges = {'veg': ['4001', '4012'],
             'vps': ['5001', '5002'],
             'vig': ['5003', '5005'],
             'fwv': ['4013', '4028'],
             'pwv': ['5006', '5009']}
   
   # Определяем колонку
   for row in range(3, 8):
       for col in range(3):
           if str(sheet[row][col].value) == '4001':
               column = col
               break


   for row in range(1, 70):
       for key, value in ranges.items():
           if str(sheet[row][column].value) in value:
               i = value.index(str(sheet[row][column].value))
               value[i] = row

   # Определяем колонки окончания диапазонов
   for column in range(40, 70):
       if type(sheet[1][column]) != MergedCell:
           end_column_letter = str(sheet[1][column - 1].coordinate)[:-1]
           break
         
   start_column_letter = 'B'
        
   veg = sheet[start_column_letter + str(ranges['veg'][0]): end_column_letter + str(ranges['veg'][1])]
   vps = sheet[start_column_letter + str(ranges['vps'][0]): end_column_letter + str(ranges['vps'][1])]
   vig = sheet[start_column_letter + str(ranges['vig'][0]): end_column_letter + str(ranges['vig'][1])]
   fwv = sheet[start_column_letter + str(ranges['fwv'][0]): end_column_letter + str(ranges['fwv'][1])]
   pwv = sheet[start_column_letter + str(ranges['pwv'][0]): end_column_letter + str(ranges['pwv'][1])]

   categories = [(veg, 'VEG'), (vps, 'VPS'), (vig, 'VIG'), (fwv, 'FWV'), (pwv, 'PWV')]
   return categories


# Функция определяющая правильное склонение существительного
def choose_plural(amount, declensions): 
    if 5 <= (amount % 100) <= 20:
         return f'{amount} {declensions[2]}'   
    elif amount % 10 == 1:
        return f'{amount} {declensions[0]}'
    elif 2 <= (amount % 10) < 5:
        return f'{amount} {declensions[1]}'
    else:
        return f'{amount} {declensions[2]}'     


# Функция настройки разметки первого листа excel
def prepare_first_sheet(sheet):
    # Настройка длины ячеек
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 45
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 17
    sheet.column_dimensions['E'].width = 30
    sheet.column_dimensions['F'].width = 30
    sheet.column_dimensions['G'].width = 25
    sheet.column_dimensions['H'].width = 25
    sheet.column_dimensions['I'].width = 20

    # Настройка высоты ячеек 
    for i in range(1, 1500):
        sheet.row_dimensions[i].height = 20

    # Cоздание заголовков
    sheet['A1'] = 'БАТЛЕР'
    sheet['B1'] = 'ГОСТЬ'
    sheet['C1'] = 'КАТЕГОРИЯ'
    sheet['D1'] = 'ВИЛЛА'
    sheet['E1'] = 'ТАРИФ'
    sheet['F1'] = 'ЗАЕЗД'
    sheet['G1'] = 'ВЫЕЗД'
    sheet['H1'] = 'КОЛИЧЕСТВО СУТОК'
    sheet['I1'] = 'CМЕННОСТЬ'

    # Форматирование и выравнивание заголовков
    for column in range(9):
        sheet[1][column].font = Font(name='Arial', bold=True)
        sheet[1][column].alignment = Alignment(horizontal="center", vertical="center")


# Функция настройки разметки второго листа excel
def prepare_second_table(sheet1):
    season, start_period, end_period = detrmine_season()
    # Объединение ячеек
    sheet1.merge_cells('A1:H1')
    sheet1.merge_cells('I1:T1')
    sheet1.merge_cells('A2:A3')
    sheet1.merge_cells('B2:H2')
    sheet1.merge_cells('I2:L2')
    sheet1.merge_cells('M2:N2')
    sheet1.merge_cells('O2:O3')
    sheet1.merge_cells('P2:V2')

    # Вписываем названия столбцов
    sheet1['A1'] = f'Расчетный период: {start_period.strftime("%d.%m.%y")}-{end_period.strftime("%d.%m.%y")} - {season}'
    sheet1['I1'] = 'Статистика'
    sheet1['A2'] = 'Батлер'

    sheet1['B2'] = 'Количество вилл'
    sheet1['B3'] = 'Всего:'
    sheet1['C3'] = 'Дни:'
    sheet1['D3'] = 'veg'
    sheet1['E3'] = 'fwv'
    sheet1['F3'] = 'pwv'
    sheet1['G3'] = 'vps'
    sheet1['H3'] = 'vig'
    sheet1['I2'] = 'Тариф'

    sheet1['M2'] = 'Сменность'
    sheet1['M3'] = 'Один'
    sheet1['N3'] = '2/2'
    
    sheet1['O2'] = 'Коэфф.'

    sheet1['P2'] = 'Занятость'
    sheet1['P3'] = 'Статус:'
    sheet1['Q3'] = 'Дата выезда:'
    sheet1['R3'] = 'Гость:'
    sheet1['S3'] = 'Планируемые заезды:'
    sheet1['T3'] = 'Ближайший заезд:'
    sheet1['U3'] = 'Дней до заезда:'
    sheet1['V3'] = 'Гость:'

    sheet1['I3'].fill = PatternFill('solid', fgColor='FFED7D31')
    sheet1['J3'].fill = PatternFill('solid', fgColor='FF00B050')
    sheet1['K3'].fill = PatternFill('solid', fgColor='FFD0CECE') 
    sheet1['L3'].fill = PatternFill('solid', fgColor='FFFFFF00') 

    # Настраиваем ширину ячеек
    sheet1.column_dimensions['A'].width = 20
    sheet1.column_dimensions['B'].width = 8
    sheet1.column_dimensions['C'].width = 8
    for col in range(3, 13):
        letter = (sheet1[1][col].coordinate)[0]
        sheet1.column_dimensions[letter].width = 5
    sheet1.column_dimensions['M'].width = 8
    sheet1.column_dimensions['N'].width = 8
    sheet1.column_dimensions['O'].width = 10
    sheet1.column_dimensions['P'].width = 15
    sheet1.column_dimensions['Q'].width = 17
    sheet1.column_dimensions['R'].width = 22
    sheet1.column_dimensions['S'].width = 38
    sheet1.column_dimensions['T'].width = 22 
    sheet1.column_dimensions['U'].width = 20
    sheet1.column_dimensions['V'].width = 20   

    # Настройка высоты ячеек 
    for i in range(1, 310):
        sheet1.row_dimensions[i].height = 20

    # Настраиваем размер, шрифт, расположение текста шапки
    for row in range(1, 4):
        for column in range(22):
            sheet1[row][column].font = Font(name='Arial', size=12, bold=True)
            sheet1[row][column].alignment = Alignment(horizontal="center", vertical="center")

    # Настраиваем обрамление ячеек
    def set_border(ws, cell_range):
        thin = Side(border_style="thin", color="000000")
        for row in ws[cell_range]:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    set_border(sheet1, 'A1:V3') 






