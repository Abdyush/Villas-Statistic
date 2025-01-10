import os
import sys
import pickle
sys.path.append(os.getcwd())

import flet as ft
import openpyxl
from data.collection import new_month_statistic, butlers, perfect_new_month_statistic, fill_dict
from data.data_output import new_fill_first_table, new_fill_second_table
from data.processing import new_prepare_dict, new_new_count_totals
from copy_file import copy_file_to_folder

def main(page: ft.Page):
#------------------------------------------------------ПЕРВАЯ СТРАНИЦА---------------------------------------------------------------------
    def first_page():
        page.clean()
        page.add(
            ft.Column([ft.Text("Это первая страница"), ft.ElevatedButton("Перейти на вторую страницу", on_click=switch_to_second_page)]))

        # Переменные для хранения путей к файлам
        file_path_1 = ""
        file_path_2 = ""
        file_name_1 = ''
        file_name_2 = ''
        paths = {'one': '',
                 'two': ''}

        #---------------------------# Создаем кнопку выбора первого файла
        def select_first_file(e: ft.FilePickerResultEvent):
            page.add(filepicker_1)
            filepicker_1.pick_files("Выберите файл с Occupancy за прошлый год", allowed_extensions=["xls", "xlsx", "xlsm", "xltx", "xltm"]) 
            
            
        # Создаем кнопку выбора второго файла
        def select_second_file(e: ft.FilePickerResultEvent):
            page.add(filepicker_2)
            filepicker_2.pick_files("Выберите файл с Occupancy за текущий год", allowed_extensions=["xls", "xlsx", "xlsm", "xltx", "xltm"])
        
        # Функция получения пути первого файла
        def return_first_file(e: ft.FilePickerResultEvent): 
            nonlocal file_path_1  # Используем global для изменения переменной в области видимости
            nonlocal file_name_1
            if e.files:  # Проверяем, что файл был выбран
                file_path_1 = e.files[0].path
                file_name_1 = e.files[0].name
                file_path_label_1.value = f"Выбранный файл 2: {file_name_1}"
                confirm_button_1.visible = True  # Показываем кнопку подтверждения
                confirm_button_1.disabled = False  
                cancel_button_1.visible = True  # Показываем кнопку отмены
                select_button_1.disabled = True  # Делаем кнопку выбора файла недоступной
            else:
                file_path_label_1.value = "Файл не выбран."
                confirm_button_1.visible = False  # Скрываем кнопку подтверждения
                cancel_button_1.visible = False  # Скрываем кнопку отмены
            page.update()

        def return_second_file(e: ft.FilePickerResultEvent): 
            nonlocal file_path_2  # Используем global для изменения переменной в области видимости
            nonlocal file_name_2
            if e.files:  # Проверяем, что файл был выбран
                file_path_2 = e.files[0].path
                file_name_2 = e.files[0].name
                file_path_label_2.value = f"Выбранный файл 2: {file_name_2}"
                confirm_button_2.visible = True  # Показываем кнопку подтверждения
                confirm_button_2.disabled = False  
                cancel_button_2.visible = True  # Показываем кнопку отмены
                select_button_2.disabled = True
            else:
                file_path_label_2.value = "Файл не выбран."
                confirm_button_2.visible = False  # Скрываем кнопку подтверждения
                cancel_button_2.visible = False  # Скрываем кнопку отмены
            page.update()
        
        
        def confirm_file_1(e):
            paths['one'] = file_path_1
            file_path_label_1.value = f"Файл подтвержден: {file_name_1}"
            confirm_button_1.disabled = True
            select_button_1.visible = False
            if paths['one'] != '' and paths['two'] != '':
                btn.disabled = False
            page.update()  
        
        def confirm_file_2(e):
            paths['two'] = file_path_2
            file_path_label_2.value = f"Файл подтвержден: {file_name_2}"
            confirm_button_2.disabled = True
            select_button_2.visible = False
            if paths['one'] != '' and paths['two'] != '':
                btn.disabled = False
            page.update()  

        def cancel_file_1(e):
            nonlocal file_path_1  # Объявляем переменную как глобальную
            nonlocal file_name_1
            file_path_1 = None  # Сбрасываем путь к файлу
            file_name_1 = ''
            paths['one'] = file_path_1
            file_path_label_1.value = "Путь выбранного файла 1"  # Сбрасываем текст метки
            confirm_button_1.visible = False  # Скрываем кнопку подтверждения
            cancel_button_1.visible = False  # Скрываем кнопку отмены
            select_button_1.visible = True
            select_button_1.disabled = False
            if paths['one'] == None or paths['two'] == None:
                btn.disabled = True
            page.update()

        def cancel_file_2(e):
            nonlocal file_path_2  # Объявляем переменную как глобальную
            nonlocal file_name_2
            file_path_2 = None  # Сбрасываем путь к файлу
            file_name_2 = ''
            paths['two'] = file_path_2
            file_path_label_2.value = "Путь выбранного файла 2"  # Сбрасываем текст метки
            confirm_button_2.visible = False  # Скрываем кнопку подтверждения
            cancel_button_2.visible = False  # Скрываем кнопку отмены
            select_button_2.visible = True
            select_button_2.disabled = False
            if paths['one'] == None or paths['two'] == None:
                btn.disabled = True
            page.update()
    
        def main_work(e, last_year, this_year):
            target_folder = 'distribution'
            try:
                # Путь к файлу эксель с распределением вилл 
                #FILE_PATH = '/home/user/Рабочий стол/StudyProject/villas/distribution/Villas occupancy 2023.xlsx' 
                FILE_PATH = copy_file_to_folder(last_year, target_folder)
                

                # Путь к файлу эксель с распределением вилл 
                #FILE_24_PATH = '/home/user/Рабочий стол/StudyProject/villas/distribution/Villas occupancy 2024.xlsx'
                FILE_24_PATH = copy_file_to_folder(this_year, target_folder)

                # Открываем файл
                file = openpyxl.open(FILE_PATH)
                file2 = openpyxl.open(FILE_24_PATH)


                # определяем список со всеми доступными страницами файла
                sheets = file.worksheets
                sheets2 = file2.worksheets

                list_sheet = [sheets, sheets2]

                other_comments = []

                # Добавляем батлеров из файла в словарь
                fill_dict(butlers)

                #Перебирая листы файла за каждый месяц, добавляем каждый лист в функцию получения статистики, таким образом, получаем статистику за год
                for sheets in list_sheet:
                    for sheet in sheets:
                        perfect_new_month_statistic(sheets=sheets, sheet=sheet, sheet_index=sheets.index(sheet), 
                                                    sheets_index=list_sheet.index(sheets))

                new_prepare_dict(butlers)

                # Закрываем файлы
                file.close()  
                file2.close()

                # Удаляем файлы
                os.remove(FILE_PATH)
                os.remove(FILE_24_PATH)

                generaly_dictionary = new_new_count_totals(butlers)

                # open a pickle file
                filename = 'selected_butlers.pk'

                with open(filename, 'wb') as fi:
                    # dump your data into the file
                    pickle.dump([], fi)

                # Путь к файлу со статистикой
                STATISTIC_FILE_PATH = '/home/user/Рабочий стол/StudyProject/villas/Statistic/Butlers_statistic.xlsx'

                # Открываем файл cо статистикой
                file_stat = openpyxl.load_workbook(STATISTIC_FILE_PATH)

                # Удаляем все листы из файла
                all_sheet = file_stat.sheetnames
                for sh in all_sheet:
                    file_stat.remove(file_stat[sh])

                # Создаем листы "Все виллы", "Статистика"
                sheet = file_stat.create_sheet("Все виллы")
                sheet1 = file_stat.create_sheet("Статистика")

                # Заполняем первую таблицу всеми данными по виллам
                new_fill_first_table(sheet, dict(sorted(butlers.items())))

                # Заполняем вторую таблицу статистикой
                new_fill_second_table(sheet1, generaly_dictionary)

                # Выводим в терминал "Готово!", сохраняем и закрываем файл
                print('Готово!')
                file_stat.save(STATISTIC_FILE_PATH)
                file_stat.close()

                result = ft.Text(value="Готово!", color="green")
                page.add(result)
                error_message.value = ""
            except Exception as ex:
                # Отображаем сообщение об ошибке
                error_message.value = f"Ошибка: {str(ex)}"
            # Обновляем экран
            page.update()

        def start_programm(e):
            if len(paths) == 2:
                file_1 = paths['one']
                file_2 = paths['two']
                main_work(e, file_1, file_2)
            else:
                print('Подтвердите пути файлов')



        #----------------------------------Кнопки и ряды---------------------------------------------------------

        # Создаем текстовый элемент для отображения ошибок
        error_message = ft.Text(value="", color="red", size=16)

        # Создаем ряды и поля путей файлов
        row_filepicker_1 = ft.Row(vertical_alignment="center")
        file_path_label_1 = ft.Text(value="Путь выбранного файла 1", expand=1)

        row_filepicker_2 = ft.Row(vertical_alignment="center")
        file_path_label_2 = ft.Text(value="Путь выбранного файла 2", expand=1)

        # Создаем filpickers - элементы управления для выбора файла
        filepicker_1 = ft.FilePicker(on_result=return_first_file)
        select_button_1 = ft.ElevatedButton(text="Выберите файл с Occupancy за прошлый год", on_click=select_first_file)
        row_filepicker_1.controls.append(select_button_1)
        
        filepicker_2 = ft.FilePicker(on_result=return_second_file)
        select_button_2 = ft.ElevatedButton(text="Выберите файл с Occupancy за текущий год", on_click=select_second_file)
        row_filepicker_2.controls.append(select_button_2)

        # Добавляем пути файлов в filepicker
        row_filepicker_1.controls.append(file_path_label_1)
        row_filepicker_2.controls.append(file_path_label_2)

        # Кнопки подтверждения
        confirm_button_1 = ft.ElevatedButton(text="Подтвердить файл 1", on_click=confirm_file_1, visible=False)
        confirm_button_2 = ft.ElevatedButton(text="Подтвердить файл 2", on_click=confirm_file_2, visible=False)

         # Кнопки отмены файла
        cancel_button_1 = ft.ElevatedButton(text="Отменить выбор файла 1", on_click=cancel_file_1, visible=False)
        cancel_button_2 = ft.ElevatedButton(text="Отменить выбор файла 1", on_click=cancel_file_2, visible=False)

        # Добавляем кнопки подтверждения в ряды
        row_filepicker_1.controls.append(cancel_button_1)
        row_filepicker_2.controls.append(cancel_button_2)

        # Добавляем кнопки подтверждения в ряды
        row_filepicker_1.controls.append(confirm_button_1)
        row_filepicker_2.controls.append(confirm_button_2)

        btn = ft.FilledButton(text='Запуск', width=300, height=45, on_click=start_programm, disabled=True)
        page.add(btn, error_message)

        # Добавляем ряды файлпикеров на страницу
        page.add(row_filepicker_1)
        page.add(row_filepicker_2)
        page.update()

        page.update()


#------------------------------------------------------ВТОРАЯ СТРАНИЦА---------------------------------------------------------------------
    def second_page():
        page.clean()  # Очищаем страницу
        page.add(
            ft.Column([ft.Text("Это вторая страница"), ft.ElevatedButton("Вернуться на первую страницу", on_click=switch_to_first_page)]))
        
            # open a pickle file
        filename1 = 'all_butlers.pk'
        filename2 = 'selected_butlers.pk'

        # load your data back to memory when you need it
        with open(filename1, 'rb') as fi:
            all_butlers = pickle.load(fi)

        first_shift = all_butlers[0]
        second_shift = all_butlers[1]


        # Создаем контейнеры для сотрудников
        first_shift_list = ft.Column()
        second_shift_list = ft.Column()

        def populate_employee_lists(first_shift, second_shift):
            # Заполняем колонки начальными данными
            first_shift_list.controls.clear()
            second_shift_list.controls.clear()
            
            for name in first_shift:
                first_shift_list.controls.append(create_employee_row(name, first_shift_list))
            
            for name in second_shift:
                second_shift_list.controls.append(create_employee_row(name, second_shift_list))
            
            page.update()

        def add_employee(shift_column):
            # Добавляем нового сотрудника в выбранную смену
            employee_name = employee_input.value.strip()
            if employee_name:
                shift_column.controls.append(create_employee_row(employee_name, shift_column))
                employee_input.value = ""
                page.update()

        def create_employee_row(name, shift_column):
            # Создаем строку с именем сотрудника и чекбоксом
            checkbox = ft.Checkbox(label="", value=False)  # Чекбокс для выбора
            row = ft.Row(
                controls=[
                    ft.TextField(value=name, width=200, on_change=lambda e: update_employee_name(e, row)),
                    checkbox,
                    ft.IconButton(ft.Icons.DELETE, on_click=lambda e: delete_employee(row, shift_column))
                ],
                alignment=ft.MainAxisAlignment.SPACE_BETWEEN
            )
            return row

        def delete_employee(row, shift_column):
            # Удаляем сотрудника из списка
            shift_column.controls.remove(row)
            page.update()

        def update_employee_name(e, row):
            # Обновляем имя сотрудника в строке
            new_name = e.control.value.strip()
            if new_name:
                row.controls[0].value = new_name
                page.update()

        def confirm_changes(e):
            # Подтверждаем изменения и возвращаем обновленные списки
            updated_first_shift = [row.controls[0].value for row in first_shift_list.controls]
            updated_second_shift = [row.controls[0].value for row in second_shift_list.controls]
            all_but = [updated_first_shift, updated_second_shift]
            with open(filename1, 'wb') as fi:
            # dump your data into the file
                pickle.dump(all_but, fi)
            print("Обновленный список первой смены:", updated_first_shift)
            print("Обновленный список второй смены:", updated_second_shift)

        def confirm_selection(e):
            # Подтверждаем выбор сотрудников с галочками
            selected_first_shift = []
            selected_second_shift = []
            
            for row in first_shift_list.controls:
                checkbox = row.controls[1]  # Получаем чекбокс из строки
                if checkbox.value:
                    selected_first_shift.append(row.controls[0].value)  # Добавляем имя в список если чекбокс отмечен
            
            for row in second_shift_list.controls:
                checkbox = row.controls[1]  # Получаем чекбокс из строки
                if checkbox.value:
                    selected_second_shift.append(row.controls[0].value)  # Добавляем имя в список если чекбокс отмечен
            sel_but = [selected_first_shift, selected_second_shift]
            with open(filename2, 'wb') as fi:
            # dump your data into the file
                pickle.dump(sel_but, fi)
            print("Выбранные сотрудники:", selected_first_shift) 
            print("Выбранные сотрудники:", selected_second_shift)


        # Поле для ввода фамилии сотрудника
        employee_input = ft.TextField(label="Фамилия сотрудника", width=300)
        
        # Кнопки добавления для каждой смены
        add_first_shift_button = ft.IconButton(ft.Icons.ADD, on_click=lambda e: add_employee(first_shift_list))
        add_second_shift_button = ft.IconButton(ft.Icons.ADD, on_click=lambda e: add_employee(second_shift_list))
        
        # Кнопка подтверждения изменений
        confirm_button = ft.Button(text="Подтвердить изменения", on_click=confirm_changes)

        # Кнопка подтверждения выбора
        confirm_selection_button = ft.Button(text="Статистика отдельных батлеров", 
                                            tooltip="Для просмотра статистики отдельных батлеров, поставьте галочки напротив их фамилий", 
                                            on_click=confirm_selection)

        # Заполняем начальные данные
        populate_employee_lists(first_shift, second_shift)

        # Добавляем элементы на страницу
        page.add(
            ft.Row(controls=[
                ft.Column(controls=[ft.Text("Первая смена"), first_shift_list, employee_input, add_first_shift_button]),
                ft.Column(controls=[ft.Text("Вторая смена"), second_shift_list, add_second_shift_button])
            ]),
            confirm_button,
            confirm_selection_button
        )
        
    # Функция для переключения на первую страницу
    def switch_to_first_page(e):
        first_page()

    # Функция для переключения на вторую страницу
    def switch_to_second_page(e):
        second_page()

    # Запускаем приложение с первой страницы
    first_page()

ft.app(main)