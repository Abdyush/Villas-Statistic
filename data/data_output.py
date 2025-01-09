from openpyxl.styles import Font, PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

from src.background import prepare_first_sheet, prepare_second_table, choose_plural


#------------------------------------ ВЫВОД ДАННЫХ В ECXEL ---------------------------------------------------------

def new_fill_first_table(sheet, butlers):
    # Подготавливаем таблицу
    prepare_first_sheet(sheet)
    
    # Заполняем словарь
    row = 2
    col = 3
    thin = Side(border_style="thin", color="000000")
    for butler, villas in butlers.items():
        if villas == {}:
            continue
        ran = f'B{row}:I{row}'
        sheet.merge_cells(ran)
        cell = sheet.cell(row=row, column=2)
        cell.value = 'Проведенные и текущие виллы'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name='Arial', size=12, bold=True)
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)      

        for time, guest in villas.items():
            cell = sheet.cell(row=row, column=1)
            if time == 'future':
                cell.value == None
            else:
                cell.value = butler
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name='Arial', size=12, bold=True)
            
                row += 1

            if time == 'future':
                row -= 1
                ran = f'B{row}:I{row}'
                sheet.merge_cells(ran)
                cell = sheet.cell(row=row, column=2)
                cell.value = 'Запланированные виллы'
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name='Arial', size=12, bold=True)
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                row += 1

            for g, info in guest.items():
                cell = sheet.cell(row=row, column=2)
                cell.value = g
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name='Arial', size=12, bold=True)
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            
                for iterance in info:
                    if info.index(iterance) > 0:
                        row += 1
                    cell = sheet.cell(row=row, column=2)
                    cell.value = g
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.font = Font(name='Arial', size=12, bold=True)
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    col = 3
                    for param in iterance:
                        if type(param) == list:
                            cell = sheet.cell(row=row, column=col)
                            cell.value = param[0][0].strftime('%d.%m.%y')
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            cell.font = Font(name='Arial', size=12)
                            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                            col += 1
                        
                            cell = sheet.cell(row=row, column=col)
                            cell.value = param[0][1].strftime('%d.%m.%y')
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            cell.font = Font(name='Arial', size=12)
                            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                            col += 1

                            cell = sheet.cell(row=row, column=col)
                            cell.value = choose_plural(param[1][1], ("день", "дня", "дней"))
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            cell.font = Font(name='Arial', size=12)
                            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                            col += 1

                        elif iterance.index(param) == 2:
                            cell = sheet.cell(row=row, column=col)
                            cell.value = param
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            cell.font = Font(name='Arial', size=12)
                            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

                            if param == 'Открытый рынок':
                                for j in range(2, 10):
                                    sheet.cell(row, j).fill = PatternFill('solid', fgColor='FFED7D31')
                            elif param == 'Сбер':
                                for j in range(2, 10):
                                    sheet.cell(row, j).fill = PatternFill('solid', fgColor='FF00B050')
                            elif param == 'Комплиментарный тариф':
                                for j in range(2, 10):
                                    sheet.cell(row, j).fill = PatternFill('solid', fgColor='FFFFFF00')
                            elif param == 'Апгрейд':
                                for j in range(2, 10):
                                    sheet.cell(row, j).fill = PatternFill('solid', fgColor='FFD0CECE')
                            else:
                                for j in range(2, 10):
                                    sheet.cell(row, j).fill = PatternFill('solid', fgColor='00C0C0C0')       
                        
                            col += 1
            
                        else:
                            cell = sheet.cell(row=row, column=col)
                            cell.value = param
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            cell.font = Font(name='Arial', size=12)
                            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                            col += 1
                                
                row += 1
            
            row += 1
          

# Функция заполнения второй таблицы словарем адаптированная к сменам
def new_fill_second_table(sheet1, general_dictionary):
    # Настраиваем разметку таблицы
    prepare_second_table(sheet1=sheet1)

    # Заполняем таблицу статистикой
    rows = []
    for j in range(1, 22):
        sheet1.cell(4, j).fill = PatternFill('solid', fgColor='E0E0E0')
    row = 5
    col = 1
    thin = Side(border_style="thin", color="000000")
    for shift in general_dictionary.values():
        for butler, stat in shift.items():
            cell = sheet1.cell(row=row, column=col)
            cell.value = butler
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name='Arial', size=12, bold=True)
            cell.border = Border(top=thin, right=thin, bottom=thin)
            col += 1
            for lst in stat:
                for dct in lst:
                    for key, value in dct.items():
                        cell = sheet1.cell(row=row, column=col)
                        if type(value) == list:
                            value = ', '.join(value)
                        if value == 'Свободен':
                            cell.fill = PatternFill('solid', fgColor='00FF00')
                        elif value == 'С виллой':
                            cell.fill = PatternFill('solid', fgColor='FF0000')
                        cell.value = value
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.font = Font(name='Arial', size=12)
                        if col == 8 or col == 12 or col == 14 or col == 15 or col == 17 or col == 21 or col == 2 or col == 3:
                            cell.border = Border(top=thin, right=thin, bottom=thin)
                        else:
                            cell.border = Border(top=thin, bottom=thin)
                        col += 1
            row += 1
            col = 1
        
        end_range = row
        rows.append(end_range)
        row += 1

    for j in range(1, 22):
        sheet1.cell(rows[0], j).fill = PatternFill('solid', fgColor='E0E0E0')
    ran1 = f'A4:U{rows[0] - 1}'
    ran2 = f'A{rows[0]}:U{row}'
    tab = Table(displayName='Table1', ref=ran1)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    tab2 = Table(displayName='Table2', ref=ran2)


    sheet1.add_table(tab)
    sheet1.add_table(tab2)

        