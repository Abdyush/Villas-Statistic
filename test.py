import os
import sys
sys.path.append(os.getcwd())

import openpyxl
from data.collection import month_statistic, new_month_statistic, butlers
from data.data_output import fill_first_table, fill_second_table, new_fill_first_table, new_fill_second_table
from data.processing import prepare_dict, count_totals, new_prepare_dict, new_count_totals

# Путь к файлу эксель с распределением вилл 
FILE_PATH = '/home/user/Рабочий стол/StudyProject/villas/distribution/Villas occupancy 2023.xlsx' 

# Путь к файлу эксель с распределением вилл 
FILE_24_PATH = '/home/user/Рабочий стол/StudyProject/villas/distribution/Villas occupancy 2024.xlsx' 

# Открываем файл
file = openpyxl.open(FILE_PATH)
file2 = openpyxl.open(FILE_24_PATH)

# определяем список со всеми доступными страницами файла
sheets = file.worksheets
sheets2 = file2.worksheets

#print(sheets2[1]['AV12'].fill.start_color.index)

list_sheet = [sheets, sheets2]

other_comments = []

#Перебирая листы файла за каждый месяц, добавляем каждый лист в функцию получения статистики, таким образом, получаем статистику за год
for sheets in list_sheet:
    for sheet in sheets:
        new_month_statistic(sheets=sheets, sheet=sheet, sheet_index=sheets.index(sheet), sheets_index=list_sheet.index(sheets))



new_prepare_dict(butlers)



#print(butlers)

# Закрываем файл
file.close()   

generaly_dictionary = new_count_totals(butlers)

#print(generaly_dictionary)

# Путь к файлу со статистикой
STATISTIC_FILE_PATH = '/home/user/Рабочий стол/StudyProject/villas/Statistic/Butlers_statistic.xlsx'

# Открываем файл cо статистикой
file_stat = openpyxl.load_workbook(STATISTIC_FILE_PATH)

# Удаляем все листы из файла
#all_sheet = file_stat.sheetnames
#for sh in all_sheet:
    #file_stat.remove(file_stat[sh])

# Создаем листы "Все виллы", "Статистика"
sheet = file_stat.create_sheet("Все виллы")
sheet1 = file_stat.create_sheet("Статистика")

# Заполняем первую таблицу всеми данными по виллам
new_fill_first_table(sheet, dict(sorted(butlers.items())))

# Заполняем вторую таблицу статистикой
new_fill_second_table(sheet1, generaly_dictionary)

# Выводим в терминал "Готово!", сохраняем и закрываем файл
print('Готово!')
file_stat.save(STATISTIC_FILE_PATH)
file_stat.close()

